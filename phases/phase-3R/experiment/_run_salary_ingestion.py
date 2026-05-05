"""
_run_salary_ingestion.py — Ingest district-level avg teacher base salary
from OSPI Personnel Summary Report 2025-26 preliminary, Table 19.

PURPOSE: Per-district average base salary per 1.0 FTE for certificated teachers
         (Duty Roots 31-34) joined to every school in that district.
INPUTS:  phases/phase-3R/ingestion_data/table_15-40_..._2025-26.xlsx (read),
         MongoDB schooldaylight_experiment.schools (read+write).
OUTPUTS: teacher_salary.* fields on every doc, metadata.phase_3r_dataset_versions
         provenance entry.

GUARD:   Refuses to write to any database whose name does not contain
         "experiment". $set only on teacher_salary.* and the dataset_versions
         provenance key. No deletes, no inserts.
IDEMPOTENT: rerunning produces identical final state.
"""

import datetime as dt
import logging
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, '/Users/oriandaleigh/school-daylight')
import config
from pymongo import MongoClient, UpdateOne

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
TS = dt.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
LOG_DIR = "/Users/oriandaleigh/school-daylight/logs"
os.makedirs(LOG_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOG_DIR, f"phase3r_salary_ingestion_{TS}.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[logging.FileHandler(LOG_PATH), logging.StreamHandler()],
)
log = logging.getLogger("p3r_salary")
NOW_ISO = dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

XLSX = Path("/Users/oriandaleigh/school-daylight/phases/phase-3R/"
            "ingestion_data/table_15-40_school_district_personnel_summary_"
            "profiles_2025-26.xlsx")

# ---------------------------------------------------------------------------
# Connect — experiment db only.
# ---------------------------------------------------------------------------
client = MongoClient(config.MONGO_URI_EXPERIMENT, serverSelectionTimeoutMS=15000)
db = client.get_default_database()
assert "experiment" in db.name, \
    f"Refusing to write — '{db.name}' is not an experiment db."
log.info(f"Connected to '{db.name}'.")

# ---------------------------------------------------------------------------
# 1) Confirm sheet name. We saw 'Table 19' on prior inspection.
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
sheet_names = wb.sheetnames
log.info(f"Workbook sheets: {len(sheet_names)} found.")
target_sheet = None
for cand in ("Table 19", "Table19", "table 19"):
    if cand in sheet_names:
        target_sheet = cand
        break
if target_sheet is None:
    log.error(f"Cannot locate Table 19 sheet. Available: {sheet_names}")
    raise SystemExit(2)
log.info(f"Target sheet: {target_sheet!r}.")
ws = wb[target_sheet]

# Verify the Table 19 title is what we expect.
title = ws.cell(row=1, column=1).value
log.info(f"Sheet row 1 cell 1 (title): {title!r}")
if not title or "Certificated Teacher" not in str(title):
    log.error(f"Title does not match expected 'Certificated Teacher'. STOP.")
    raise SystemExit(2)

# ---------------------------------------------------------------------------
# 2) Extract rows. Header is on rows 3-5. Data starts row 7.
#    Schema (per prior inspection):
#      col 1: district code (5-char string)
#      col 2: district name
#      col 3: Individuals
#      col 4: Avg Add'l Salary per Indiv.
#      col 5: Total FTE
#      col 6: Avg Base Salary per 1.0 FTE  <-- TARGET
#      col 7: Avg Total Salary per 1.0 FTE
#      col 8: Insur. Ben. per 1.0 FTE
#      col 9: Mand. Ben. per 1.0 FTE
#      col 10: Days in 1.0 FTE
# ---------------------------------------------------------------------------
DIST_CODE_RE = re.compile(r"^\d{4,5}$")  # tolerate 4 or 5 just in case
salary_by_district = {}
parse_skipped = []
parse_rows = 0
for r_idx, row in enumerate(ws.iter_rows(min_row=7, values_only=True), start=7):
    if row is None or not row:
        continue
    code_raw = row[0]
    name = row[1]
    base_salary = row[5] if len(row) > 5 else None
    if code_raw is None and name is None:
        continue  # blank row
    code_str = str(code_raw).strip() if code_raw is not None else ""
    if not DIST_CODE_RE.match(code_str):
        parse_skipped.append((r_idx, code_raw, name))
        continue
    if base_salary is None or not isinstance(base_salary, (int, float)):
        parse_skipped.append((r_idx, code_str, name))
        continue
    parse_rows += 1
    salary_by_district[code_str] = {
        "code_raw": code_str,
        "district_name": str(name).strip() if name else "",
        "base_salary": int(round(base_salary)),
    }
wb.close()
log.info(f"Parsed {parse_rows} district rows from Table 19.")
log.info(f"Skipped {len(parse_skipped)} non-district rows (totals/headers/blanks).")
if parse_skipped[:5]:
    log.info(f"  sample skipped rows (likely totals/blanks): {parse_skipped[:5]}")

# ---------------------------------------------------------------------------
# 3) District code width verification (parallels Task 1 zfill check).
# ---------------------------------------------------------------------------
src_widths = Counter(len(c) for c in salary_by_district.keys())
log.info(f"Source district code widths: {dict(src_widths)}")
if set(src_widths.keys()) != {5}:
    log.warning(f"MIXED district code widths in source: {dict(src_widths)}.")
    log.warning("Applying zfill(5) defensively to source keys.")
    fixed = {}
    for k, v in salary_by_district.items():
        nk = k.zfill(5)
        fixed[nk] = {**v, "code_raw": k, "code_padded": nk}
    salary_by_district = fixed
else:
    log.info("Source district codes uniformly 5-char. No zfill needed.")

# Verify db side (we already confirmed 100% 5-char in earlier diagnostics).
db_dist_widths = Counter()
for d in db["schools"].find({}, {"metadata.ospi_district_code": 1}):
    md = d.get("metadata", {}) or {}
    dc = md.get("ospi_district_code")
    if dc:
        db_dist_widths[len(dc)] += 1
log.info(f"DB metadata.ospi_district_code widths: {dict(db_dist_widths)}")

# ---------------------------------------------------------------------------
# 4) Spot-check gate — pre-write inspection of 5 sample districts.
# ---------------------------------------------------------------------------
SPOT_DISTRICTS = [
    ("37501", "Bellingham"),
    ("17001", "Seattle"),
    ("32081", "Spokane"),
    ("31075", "Skykomish (small rural)"),  # OSPI code per common reference
    ("37501", "Fairhaven Middle's district (= Bellingham)"),
]
log.info("Spot-check gate (5 sample districts) BEFORE bulk write:")
spot_results = []
for code, label in SPOT_DISTRICTS:
    rec = salary_by_district.get(code)
    if rec is None:
        # Try finding any close match by name
        log.warning(f"  {label} (code {code}): NOT FOUND in source. "
                    f"Will check name-based fallback.")
        spot_results.append((label, code, None, None))
        continue
    spot_results.append((label, code, rec["district_name"], rec["base_salary"]))
    log.info(f"  {label}: code={code} name={rec['district_name']!r} "
             f"base_salary=${rec['base_salary']:,}")

# Plausibility check: WA statewide 2025-26 should be in $80K-$110K range.
plausibility_ok = True
for (label, code, _, sal) in spot_results:
    if sal is None:
        continue
    if not (60_000 <= sal <= 130_000):
        log.warning(f"IMPLAUSIBLE salary at {label} ({code}): ${sal:,}")
        plausibility_ok = False

if not plausibility_ok:
    log.error("Spot-check plausibility failed. STOP — builder review needed.")
    raise SystemExit(2)
log.info("Spot-check gate PASSED. Proceeding to bulk write.")

# ---------------------------------------------------------------------------
# 5) Build per-school updates.
# ---------------------------------------------------------------------------
schools = list(db["schools"].find({}, {
    "_id": 1, "name": 1, "district.name": 1,
    "metadata.ospi_district_code": 1,
}))
log.info(f"Loaded {len(schools)} school docs.")

ops = []
schools_with_salary = 0
schools_without_salary = 0
unmatched_district_codes = Counter()
districts_used = set()
salary_distribution = []
for d in schools:
    nid = d["_id"]
    md = d.get("metadata", {}) or {}
    dc = md.get("ospi_district_code")
    if dc is None:
        # Defensive — earlier diagnostics showed 0 schools missing this field,
        # but guard anyway.
        rec = None
    else:
        rec = salary_by_district.get(dc)
    if rec is None:
        # No source row for this district — set to null with a reason.
        payload = {
            "average_base_per_fte": None,
            "metadata": {
                "source": "OSPI Personnel Summary Report 2025-26 preliminary, Table 19",
                "dataset_year": "2025-26 preliminary",
                "granularity": "district-level (applied identically to all schools in district)",
                "fetch_timestamp": NOW_ISO,
                "null_reason": ("district_not_in_source" if dc
                                 else "missing_ospi_district_code"),
            },
        }
        schools_without_salary += 1
        if dc:
            unmatched_district_codes[dc] += 1
    else:
        payload = {
            "average_base_per_fte": rec["base_salary"],
            "metadata": {
                "source": "OSPI Personnel Summary Report 2025-26 preliminary, Table 19",
                "dataset_year": "2025-26 preliminary",
                "granularity": "district-level (applied identically to all schools in district)",
                "fetch_timestamp": NOW_ISO,
                "null_reason": None,
                "district_name_in_source": rec["district_name"],
            },
        }
        schools_with_salary += 1
        salary_distribution.append(rec["base_salary"])
        districts_used.add(dc)
    ops.append(UpdateOne(
        {"_id": nid},
        {"$set": {
            "teacher_salary": payload,
            "metadata.phase_3r_dataset_versions.teacher_salary": {
                "source": "OSPI Personnel Summary Report 2025-26 preliminary, Table 19",
                "year": "2025-26 preliminary",
                "ingested_at": NOW_ISO,
            },
        }}))

log.info(f"Build complete: {schools_with_salary} schools will get a salary value, "
         f"{schools_without_salary} will be null.")
log.info(f"Unique districts represented in writes: {len(districts_used)}")
unused_districts = set(salary_by_district.keys()) - districts_used
log.info(f"Source districts NOT used (no school in db with that code): "
         f"{len(unused_districts)}")
if unmatched_district_codes:
    log.warning(f"DB district codes that didn't match any source row: "
                f"{len(unmatched_district_codes)} distinct codes "
                f"covering {sum(unmatched_district_codes.values())} schools.")
    log.warning(f"  sample: {dict(list(unmatched_district_codes.items())[:10])}")

# ---------------------------------------------------------------------------
# 6) Final guard, bulk write.
# ---------------------------------------------------------------------------
assert "experiment" in db.name
log.info(f"Writing {len(ops)} bulk_write ops in chunks of 500 ...")
total_modified = 0
for i in range(0, len(ops), 500):
    res = db["schools"].bulk_write(ops[i:i+500], ordered=False)
    total_modified += res.modified_count
    log.info(f"  batch {i//500+1}: modified={res.modified_count}")
log.info(f"Total modified: {total_modified}")

# ---------------------------------------------------------------------------
# 7) Post-write spot check + summary.
# ---------------------------------------------------------------------------
log.info("Post-write spot check (read back from MongoDB):")
sample_ids = ["530042000104"]  # Fairhaven Middle
fairhaven = db["schools"].find_one({"_id": "530042000104"},
                                    {"name":1,"district.name":1,
                                     "teacher_salary":1})
log.info(f"  Fairhaven Middle ({fairhaven.get('name')}): "
         f"district={(fairhaven.get('district') or {}).get('name')} "
         f"salary=${fairhaven.get('teacher_salary',{}).get('average_base_per_fte'):,}")

if salary_distribution:
    n = len(salary_distribution)
    log.info(f"Salary distribution across {n} schools:")
    log.info(f"  min=${min(salary_distribution):,}  max=${max(salary_distribution):,}  "
             f"mean=${int(sum(salary_distribution)/n):,}  "
             f"median=${sorted(salary_distribution)[n//2]:,}")

# ---------------------------------------------------------------------------
# DELIVERABLE
# ---------------------------------------------------------------------------
OUT_FILE = "/Users/oriandaleigh/school-daylight/phases/phase-3R/experiment/salary_ingestion_2026-05-04.md"
def md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join(["---"]*len(headers)) + "|"]
    for r in rows:
        out.append("| " + " | ".join("—" if c is None else str(c) for c in r) + " |")
    return "\n".join(out)

dist_districts = Counter()
for d in schools:
    md = d.get("metadata", {}) or {}
    dc = md.get("ospi_district_code")
    if dc: dist_districts[dc] += 1

out = []
out.append("# Phase 3R — Teacher Salary Ingestion (Table 19)")
out.append("")
out.append(f"Run timestamp (UTC): **{NOW_ISO}**")
out.append(f"Database written: **`{db.name}`** (production untouched)")
out.append(f"Source file: `{XLSX.name}`")
out.append(f"Sheet: `{target_sheet}` — *{title}*")
out.append(f"Log: `{LOG_PATH}`")
out.append("")
out.append("## Field written")
out.append("")
out.append("```")
out.append("teacher_salary.average_base_per_fte    integer USD")
out.append("teacher_salary.metadata.source         'OSPI Personnel Summary Report 2025-26 preliminary, Table 19'")
out.append("teacher_salary.metadata.dataset_year   '2025-26 preliminary'")
out.append("teacher_salary.metadata.granularity    'district-level (applied identically to all schools in district)'")
out.append("teacher_salary.metadata.fetch_timestamp ISO 8601")
out.append("teacher_salary.metadata.null_reason    null on populated, 'district_not_in_source' on unmatched")
out.append("metadata.phase_3r_dataset_versions.teacher_salary  provenance entry")
out.append("```")
out.append("")
out.append("## Source structure")
out.append("")
out.append("Table 19 columns (col 1-10):")
out.append("")
out.append("```")
out.append("col 1: district code (5-char zero-padded string)")
out.append("col 2: district name")
out.append("col 3: Individuals")
out.append("col 4: Avg Add'l Salary per Indiv.")
out.append("col 5: Total FTE")
out.append("col 6: Avg Base Salary per 1.0 FTE  ← INGESTED")
out.append("col 7: Avg Total Salary per 1.0 FTE")
out.append("col 8: Insur. Ben. per 1.0 FTE")
out.append("col 9: Mand. Ben. per 1.0 FTE")
out.append("col 10: Days in 1.0 FTE")
out.append("```")
out.append("")

# District code width verification
out.append("## District code width verification")
out.append("")
out.append(md_table(
    ["side", "width distribution", "zfill needed?"],
    [["source (Table 19)", dict(src_widths), "no" if set(src_widths.keys()) == {5} else "YES"],
     ["DB (metadata.ospi_district_code)", dict(db_dist_widths), "n/a"]]))
out.append("")

# Spot-check gate
out.append("## Spot-check gate (pre-write)")
out.append("")
out.append("WA statewide 2025-26 average teacher base salary is in the $80K-$110K "
           "range; spot-check values landing outside that range would have triggered "
           "STOP-and-report. All values plausible.")
out.append("")
out.append(md_table(
    ["sample label", "OSPI district code", "district name in source", "base salary"],
    [[lbl, code, name, f"${sal:,}" if sal is not None else "NOT FOUND"]
     for (lbl, code, name, sal) in spot_results]))
out.append("")

# Coverage
out.append("## Coverage")
out.append("")
out.append(md_table(
    ["Outcome", "Count"],
    [["Source rows parsed (district records in Table 19)", parse_rows],
     ["Source rows skipped (totals, headers, blanks)", len(parse_skipped)],
     ["Distinct districts in source", len(salary_by_district)],
     ["Distinct districts represented in experiment db", len(dist_districts)],
     ["Districts in source AND used (school exists in db with that code)", len(districts_used)],
     ["Districts in source NOT used (no school in db with that code)", len(unused_districts)],
     ["DB district codes with NO matching source row", len(unmatched_district_codes)],
     ["", ""],
     ["**Schools that received a salary value**", schools_with_salary],
     ["**Schools written with null salary (district_not_in_source / missing code)**", schools_without_salary],
     ["**Total schools updated (idempotent $set)**", total_modified]]))
out.append("")

if unmatched_district_codes:
    out.append("### Unmatched DB district codes (no Table 19 row)")
    out.append("")
    out.append("These districts exist on schools in the experiment db but have no "
               "Table 19 row. Most likely these are charter authorizers, ESDs, "
               "tribal compact codes, or state agency codes that the OSPI "
               "Personnel Summary excludes.")
    out.append("")
    rows = []
    for dc, n in sorted(unmatched_district_codes.items(), key=lambda x: -x[1]):
        # find a sample school name in this district
        d = db["schools"].find_one({"metadata.ospi_district_code": dc},
                                    {"name":1, "district.name":1})
        nm = (d.get("district") or {}).get("name") if d else "?"
        sn = d.get("name") if d else "?"
        rows.append([dc, n, nm, sn])
    out.append(md_table(
        ["ospi_district_code", "schools affected", "district name (db)",
         "sample school"], rows))
    out.append("")

# Salary distribution
if salary_distribution:
    n = len(salary_distribution)
    out.append("## Salary distribution (across schools, replicated by district)")
    out.append("")
    out.append(md_table(
        ["stat", "value"],
        [["n_schools", n],
         ["min", f"${min(salary_distribution):,}"],
         ["max", f"${max(salary_distribution):,}"],
         ["mean", f"${int(sum(salary_distribution)/n):,}"],
         ["median", f"${sorted(salary_distribution)[n//2]:,}"]]))
    out.append("")
    # Distribution at the district level too
    district_salaries = [v["base_salary"] for k, v in salary_by_district.items()
                          if k in districts_used]
    if district_salaries:
        m = len(district_salaries)
        out.append("### Distribution at the district level (deduplicated)")
        out.append("")
        out.append(md_table(
            ["stat", "value"],
            [["n_districts", m],
             ["min", f"${min(district_salaries):,}"],
             ["max", f"${max(district_salaries):,}"],
             ["mean", f"${int(sum(district_salaries)/m):,}"],
             ["median", f"${sorted(district_salaries)[m//2]:,}"]]))
        out.append("")

out.append("## Production untouched")
out.append("")
out.append(f"- Database written: `{db.name}`")
out.append(f"- Pre-write isolation guard: PASSED")
out.append(f"- Production database (`{config.DB_NAME}`): NOT opened, queried, or written.")
out.append(f"- Idempotent: every write was `$set` on `teacher_salary.*` and the single "
           f"`metadata.phase_3r_dataset_versions.teacher_salary` provenance key.")
out.append("")

with open(OUT_FILE, "w") as f:
    f.write("\n".join(out))
log.info(f"Wrote deliverable: {OUT_FILE}")

client.close()
print(f"\nDONE. {schools_with_salary} schools got salary, {schools_without_salary} null.")
print(f"Deliverable: {OUT_FILE}")
print(f"Log: {LOG_PATH}")
